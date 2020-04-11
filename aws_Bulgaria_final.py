#!/usr/bin/env python
# coding: utf-8

import pandas as pd
from openpyxl import load_workbook

import numpy as np

import schedule
import time

from datetime import datetime

import plotly
import plotly.graph_objects as go
import plotly.express as px

import chart_studio

import dash
import dash_html_components as html
import dash_core_components as dcc

import cufflinks as cf

chart_studio.tools.set_credentials_file(username = 'thedjamba', api_key = '*****************')
url = "http://www.mh.government.bg/bg/informaciya-za-grazhdani/potvrdeni-sluchai-na-koronavirus-na-teritoriyata-na-r-blgariya/"

# read daily figures from url
def get_today():
    today = pd.read_html(url)[0]
    today = pd.pivot_table(today, values = 1, columns = 0)
    today.columns = ['възстановени', 'смъртни', 'потвърдени']
    today['Дата'], today['Държава'], today['събитие'] = [datetime.today().strftime('%Y-%m-%d'), 'България', np.nan]
    today = today[['Дата', 'Държава', 'потвърдени', 'смъртни', 'възстановени', 'събитие']]
    return(today)

# read excel
def get_history():
    bg = pd.read_excel("Bulgaria.xlsx")
    bg['Дата'] = bg['Дата'].dt.date
    bg['болни'] = bg['потвърдени'] - bg['смъртни'] - bg['възстановени']
    bg['Дата'] = bg['Дата'].apply(lambda x: x.strftime('%Y-%m-%d'))
    return(bg)

# check if date is the same
if get_today().iloc[0]['Дата'] == get_history().iloc[-1]['Дата']:
    # if data is the same pass
    if  (get_today().iloc[0]['потвърдени'],
         get_today().iloc[0]['смъртни'],
         get_today().iloc[0]['възстановени']) == (get_history().iloc[-1]['потвърдени'],
                                                  get_history().iloc[-1]['смъртни'],
                                                  get_history().iloc[-1]['възстановени']):
        print('pass')
        bg = get_history()

        # if data is different update excel    
    else:
        book = load_workbook('Bulgaria.xlsx')
        writer = pd.ExcelWriter('Bulgaria.xlsx', engine='openpyxl')
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        get_today().to_excel(writer, startrow = (writer.sheets['Sheet1'].max_row - 1), index = False, header = False)
        writer.save()
        bg = get_history()
        print('updated')
else:
    # if new date, add new row to excel
    book = load_workbook('Bulgaria.xlsx')
    writer = pd.ExcelWriter('Bulgaria.xlsx', engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    get_today().to_excel(writer, startrow = writer.sheets['Sheet1'].max_row, index = False, header = False)
    writer.save()
    bg = get_history()
    print('written')

# history graph
trace0 = go.Scatter (x = get_history()['Дата'], y = get_history()['болни'], name = 'болни', line = dict(width = 7))
trace1 = go.Scatter (x = get_history()['Дата'], y = get_history()['смъртни'], name = 'смъртни', line = dict(width = 7))
trace2 = go.Scatter (x = get_history()['Дата'], y = get_history()['възстановени'], name = 'възстановени', line = dict(width = 7))
data = [trace0, trace1, trace2]
layout = dict(title= 'COVID19 Разпространение в България', xaxis = dict(), yaxis = dict())
fig= dict(data = data, layout = layout)

# animation graph
bg_melt = get_history().melt(id_vars='Дата', value_vars=['болни', 'смъртни', 'възстановени'])
bg_melt.columns = ['Дата', 'Случаи', 'Брой']

# dynamic  limit for animated graph
maxi = get_history()['потвърдени'].max()
def roundup(x):
    return x if x % 100 == 0 else x + 100 - x % 100

fig1 = px.bar(bg_melt, x="Случаи", y="Брой", color="Случаи",
             animation_frame="Дата", range_y=[0, roundup(maxi)])

# create app
app = dash.Dash(__name__) 
app.title = 'COVID19 Bulgaria'
server = app.server

# add two charts
app.layout = html.Div([
    html.Div([
        html.Div([
            #html.H3('Column 1'),
            dcc.Graph(figure= dict(data = data, layout = layout)),
        ], className="six columns"),

        html.Div([
            #html.H3('Column 2'),
            dcc.Graph(figure = px.bar(bg_melt, x="Случаи", y="Брой", color="Случаи", animation_frame="Дата", range_y=[0,roundup(maxi)]))
        ], className="six columns"),
    ], className="row")
])

if __name__ == '__main__':
    server.run(debug=False)
